#xDemo = True
xDemo = False

import numpy as np
import openpyxl as pyxl
import os
import sys
import pandas
from pathlib import Path
from shutil import copyfile

from tools import AddStyles
from tools import AddTableSite
from tools import ConstructPersonDrop
from tools import ConstructResultSheet
from tools import ConstructSampleSheet
from tools import ConstructLinks
from tools import CopyData
from tools import GetMetaData
from tools import Protect
import datetime

if xDemo:
    folderLocal = './demo/'
else:
    folderLocal = './production/'
    
folderMeta = folderLocal+'meta/'
folderInput = folderLocal+'local_input/'
folderOutput = folderLocal+'local_output/'
metafile = 'MetaData.xlsx'

strPW = 'openpyxl'
strPWlab = 'SuperDaten'

# =======================================
#    1. Collect meta-information
# =======================================

metapath = os.path.join(Path(folderMeta),Path(metafile))
dfSITE, dfVAR, lstSampleMethod, lstSampleFilter = GetMetaData(metapath)

dfProject = pandas.read_excel(io =metapath, engine="openpyxl", sheet_name ='file',header=7,dtype='str',keep_default_na=False).dropna(axis=0,how='all')
dfProject.sort_values(by=['Year', 'Project'],inplace=True)

dfCentral = pandas.read_excel(io =metapath,\
                              engine="openpyxl",\
                              sheet_name ='file',\
                              header=3,\
                              nrows=2,\
                              dtype='str',\
                              keep_default_na=False).dropna(axis=0,how='all')
folderArchive = Path(dfCentral.Folder[1])
folderCentral = Path(dfCentral.Folder[0])

def RowFunc(v):
    try:
        iRow = wbOld.sheetnames.index(v)
    except:
        iRow = None
    return iRow

    
# =======================================
#    2. Create files
# =======================================

years = sorted(dfProject.Year.unique())

for i, year in enumerate(years):
    
    print('Year: '+str(year))
    
    filenameSample = 'LabSamples'+str(year)+'.xlsx'
    filepathSampleRemote = os.path.join(Path(folderCentral),filenameSample)
    filepathSampleLocalOld =  os.path.join(Path(folderInput),filenameSample)
    filepathSampleLocal =  os.path.join(Path(folderOutput),filenameSample)
    
    dfProjectYear = dfProject[dfProject.Year==str(year)].copy()
    
    filenameOld =  os.path.join(Path(folderInput),filenameSample)
    if os.path.isfile(filenameOld):
        wbOld = pyxl.load_workbook(filenameOld) #Add file name
        dfProjectYear['xAlreadyPresent'] = dfProjectYear.Project.isin(wbOld.sheetnames)
        
        dfProjectYear['order'] = dfProjectYear['Project'].apply(lambda v: RowFunc(v))
        dfProjectYear = dfProjectYear.sort_values(['xAlreadyPresent','order'])
        dfProjectYear.drop(['order'],axis=1)
        
    if os.path.isfile(filepathSampleLocalOld):
        
        wbSampleOld = pyxl.load_workbook(filepathSampleLocalOld,data_only=True) #Add file name
        
        StrPerson = 'person'
        StrPerson_db = 'person_db'
        StrPerson_drop = 'person_drop'
        
        if (StrPerson in wbSampleOld.sheetnames) and (StrPerson_db in wbSampleOld.sheetnames):
            
            # ========================================================================================
            #    Create new LabSamplesYYYY.xlsx workbook
            # ========================================================================================
            wbSample = pyxl.Workbook()
            
            # ========================================================================================
            #    Get/create sheets 'person', 'person_db', and 'person_drop'
            # ========================================================================================
            
            # ----------------------------------------------------------------------------------------
            # Get sheet 'person' from existing LabSamplesYYYY.xlsx file and copy to new workbook
            # ----------------------------------------------------------------------------------------
            wsPerson    = wbSample.active
            wsPerson.title = StrPerson
            
            # Copy existing data
            wsPerson    = CopyData(wsPerson,\
                                   folderInput=folderInput,\
                                   filename=filenameSample,\
                                   strSheetName=wsPerson.title,\
                                   row1=0,\
                                   xCopyStyle=True)
            
            # ----------------------------------------------------------------------------------------
            # Get sheet 'person_db' from existing LabSamplesYYYY.xlsx file and copy to new workbook
            # ----------------------------------------------------------------------------------------
            
            wbSample.create_sheet(StrPerson_db)
            wsPerson_db    = wbSample[StrPerson_db]
            wsPerson_db.title = StrPerson_db
            
            # Copy existing data
            wsPerson_db    = CopyData(wsPerson_db,\
                                   folderInput=folderInput,\
                                   filename=filenameSample,\
                                   strSheetName=wsPerson_db.title,\
                                   row1=0,\
                                   xCopyStyle=True)
            Protect(wsPerson_db,nRowCorner=0,nRowHead=0,nCol=0)
            wsPerson_db.protection.set_password(strPWlab)
            
            # ----------------------------------------------------------------------------------------
            # Create sheet 'person_drop' to produce actual dropdown menu
            # ----------------------------------------------------------------------------------------
            
            #wsPersonOld = wbSampleOld[StrPerson_db] #Add Sheet name
            
            wbSample.create_sheet(StrPerson_drop)
            wsPerson_drop    = wbSample[StrPerson_drop]
            wsPerson_drop.title = StrPerson_drop
            
            wsPerson_drop = ConstructPersonDrop(wsPerson_drop)
            
            # ========================================================================================
            #    Add sheets to defined possible site specifications
            # ========================================================================================
            
            wbSample = AddStyles(wbSample)
            wbSample = AddTableSite(wbSample,dfSITE=dfSITE)
            
            #print(dfProjectYear[['Year','Project']])
            
            # ========================================================================================
            #    Produce sheet for every project
            # ========================================================================================
            for ProjectInfo in dfProjectYear.itertuples(index=True, name='Pandas'):
                strSheetName = str(ProjectInfo.Project)
                
                print('  Project : '+strSheetName)
                #print(ProjectInfo)
                sys.stdout.flush()
                
                filenameResult      = 'LabResults'+str(year)+'_'+strSheetName+'.xlsx'
                filenameResultExtra = 'LabResults'+str(year)+'_'+strSheetName+'_extra.xlsx'
                filenameSampleExtra = 'LabSamples'+str(year)+'_'+strSheetName+'_extra.xlsx'
                
                if datetime.datetime.now().year<=np.float(year):
                    folder_samples = dfCentral.Folder[0]
                else:
                    folder_samples = dfCentral.Folder[1]
                    
                # Create sheet
                wbSample.create_sheet(strSheetName)
                wsSample = wbSample[strSheetName]
                wsSample.freeze_panes='C9'
                
                # Construct sheet
                if ProjectInfo.ResultFile=='True':
                    FileRes = filenameResult
                else:
                    FileRes = ""
                    
                wsSample = ConstructSampleSheet(wsSample,\
                                          ProjectInfo=ProjectInfo,\
                                          filename=FileRes,\
                                          filenameExtra="",\
                                          lstSampleMethod=lstSampleMethod,\
                                          lstSampleFilter=lstSampleFilter,\
                                          dfVAR=dfVAR,\
                                          xLab=True)
                
                # Copy existing data
                wsSample    = CopyData(wsSample,\
                                 folderInput=folderInput,\
                                 filename=filenameSample,\
                                 strSheetName=strSheetName,\
                                 row1=8,\
                                 xCopyStyle=False)
                
                
                if len( ProjectInfo.SampleID_Project)>0:
                    
                    strSheetNameExtra = "extra_"+strSheetName
                    
                # ========================================================================================
                #    Produce results file for analysis by laboratory staff
                # ========================================================================================
                
                if ProjectInfo.ResultFile=='True':
                    
                    wbResult = pyxl.Workbook()
                    wbResult = AddStyles(wbResult)
                    
                    # Create sheet
                    wsResult         = wbResult.active
                    wsResult.title   = strSheetName
                    wsResult.freeze_panes='C9'
                    
                    # Copy existing data
                    wsResult    = CopyData(wsResult,\
                                       folderInput=folderInput,\
                                       filename=filenameResult,\
                                       strSheetName=strSheetName,\
                                       row1=8,\
                                       xCopyStyle=False)
                    
                    # Construct sheet
                    wsResult = ConstructResultSheet(wsResult,\
                                                ProjectInfo=ProjectInfo,\
                                                folder=folder_samples,\
                                                filename=filenameSample,\
                                                filenameExtra="",\
                                                dfVAR=dfVAR,\
                                                xLab=True)
                    
                    # ===============================================================
                    # Create sheet with sample data from central samples sheet
                    # ===============================================================
                    wbResult.create_sheet('SamplesX')
                    wsSampleX = wbResult['SamplesX']
                    wsSampleX = ConstructLinks(wsSampleX,\
                                               strFolder=folder_samples,\
                                               strFilename=filenameSample,\
                                               strSheet=strSheetName,nRow=1007,nCol=8)
                    
                    # ===============================================================
                    # Create sheet with sample data from central samples extra sheet
                    # ===============================================================
                    if len( ProjectInfo.SampleID_Project)>0:
                        wbResult.create_sheet('SamplesY')
                        wsSampleY = wbResult['SamplesY']
                        wsSampleY = ConstructLinks(wsSampleY,\
                                                   strFolder=ProjectInfo.Folder,\
                                                   strFilename=filenameSampleExtra,\
                                                   strSheet=strSheetNameExtra,nRow=1007,nCol=8)
                        
                    # ===============================================================
                    # Store file
                    # ===============================================================
                    wbResult.security.workbookPassword = strPW
                    wbResult.security.lockStructure = True
                    filepathResultLocal =  os.path.join(Path(folderOutput),filenameResult)
                    wbResult.save(filepathResultLocal)
                
                if len( ProjectInfo.SampleID_Project)>0:
                    
                    # ========================================================================================
                    #    Produce samples file for project staff only
                    # ========================================================================================
                    
                    wbSampleExtra = pyxl.Workbook()
                    wbSampleExtra = AddStyles(wbSampleExtra)
                    
                    # ----------------------------------------------------------------------------------------
                    # Get sheet 'person' by linking to LabSamplesYYYY.xlsx file
                    # ----------------------------------------------------------------------------------------
                    
                    wsExtraPerson         = wbSampleExtra.active
                    wsExtraPerson.title   = 'person'
                    wsExtraPerson = ConstructLinks(wsExtraPerson,\
                                                   strFolder=folder_samples,\
                                                   strFilename=filenameSample,\
                                                   strSheet='person',nRow=999,nCol=4)
                    Protect(wsExtraPerson,nRowCorner=0,nRowHead=0,nCol=0)
                    
                    # ----------------------------------------------------------------------------------------
                    # Get sheet 'person_db' by linking to LabSamplesYYYY.xlsx file
                    # ----------------------------------------------------------------------------------------
                    
                    wbSampleExtra.create_sheet('person_db')
                    wsExtraPerson_db = wbSampleExtra['person_db']
                    wsExtraPerson_db.title   = 'person_db'
                    wsExtraPerson_db = ConstructLinks(wsExtraPerson_db,\
                                                   strFolder=folder_samples,\
                                                   strFilename=filenameSample,\
                                                   strSheet='person_db',nRow=999,nCol=4)
                    Protect(wsExtraPerson_db,nRowCorner=0,nRowHead=0,nCol=0)
                    
                    # ----------------------------------------------------------------------------------------
                    # Create sheet 'person_drop' to produce actual dropdown menu
                    # ----------------------------------------------------------------------------------------
            
                    wbSampleExtra.create_sheet(StrPerson_drop)
                    wsPerson_extra_drop    = wbSampleExtra[StrPerson_drop]
                    wsPerson_extra_drop.title = StrPerson_drop
                    
                    wsPerson_extra_drop = ConstructPersonDrop(wsPerson_extra_drop)
                    
                    # ----------------------------------------------------------------------------------------
                    # Add site sheets
                    # ----------------------------------------------------------------------------------------
            
                    wbSampleExtra = AddTableSite(wbSampleExtra,dfSITE=dfSITE)
            
                    # ----------------------------------------------------------------------------------------
                    # Add sheet with data
                    # ----------------------------------------------------------------------------------------
                    
                    # Create sheet
                    wbSampleExtra.create_sheet(strSheetNameExtra)
                    wsSample = wbSampleExtra[strSheetNameExtra]
                    wsSample.freeze_panes='C9'
                    
                    # Construct sheet
                    wsSample = ConstructSampleSheet(wsSample,\
                                          ProjectInfo=ProjectInfo,\
                                          filename=filenameResult,\
                                          filenameExtra=filenameResultExtra,\
                                          lstSampleMethod=lstSampleMethod,\
                                          lstSampleFilter=lstSampleFilter,\
                                          dfVAR=dfVAR,\
                                          xLab=False)
                    
                    # Copy existing data
                    wsSample    = CopyData(wsSample,\
                                           folderInput=folderInput,
                                           filename=filenameSampleExtra,
                                           strSheetName=strSheetNameExtra,
                                           row1=8,
                                           xCopyStyle=False)
                        
                    # Store file
                    wbSampleExtra.security.workbookPassword = strPW
                    wbSampleExtra.security.lockStructure = True
                    filepathSampleExtra =  os.path.join(Path(folderOutput),filenameSampleExtra)
                    wbSampleExtra.save(filepathSampleExtra)
                    
                    # ========================================================================================
                    #    Produce results file for project staff only
                    # ========================================================================================
                    
                    if ProjectInfo.ResultFile=='True':
                        
                        wbResultExtra = pyxl.Workbook()
                        wbResultExtra = AddStyles(wbResultExtra)
                        
                        # ===============================================================
                        # Create sheet with analytical results
                        # ===============================================================
                        wsResult         = wbResultExtra.active
                        wsResult.title   = strSheetNameExtra
                        wsResult.freeze_panes='C9'
                        
                        # Copy existing data
                        wsResult    = CopyData(wsResult,\
                                           folderInput=folderInput,\
                                           filename=filenameResultExtra,\
                                           strSheetName=strSheetNameExtra,\
                                           row1=8,\
                                           xCopyStyle=False)
                        
                        # Construct sheet
                        wsResult = ConstructResultSheet(wsResult,\
                                                ProjectInfo=ProjectInfo,\
                                                folder=folder_samples,\
                                                filename=filenameSample,\
                                                filenameExtra=filenameSampleExtra,\
                                                dfVAR=dfVAR,\
                                                xLab=False)
                        
                        # ===============================================================
                        # Create sheet with sample data from central samples sheet
                        # ===============================================================
                        wbResultExtra.create_sheet('SamplesX')
                        wsSampleX = wbResultExtra['SamplesX']
                        wsSampleX = ConstructLinks(wsSampleX,\
                                               strFolder=folder_samples,\
                                               strFilename=filenameSample,\
                                               strSheet=strSheetName,nRow=1007,nCol=8)
                        
                        # ===============================================================
                        # Create sheet with sample data from central samples extra sheet
                        # ===============================================================
                        
                        wbResultExtra.create_sheet('SamplesY')
                        wsSampleY = wbResultExtra['SamplesY']
                        wsSampleY = ConstructLinks(wsSampleY,\
                                                   strFolder=ProjectInfo.Folder,\
                                                   strFilename=filenameSampleExtra,\
                                                   strSheet=strSheetNameExtra,nRow=1007,nCol=8)
                        
                        # ===============================================================
                        # Store file
                        # ===============================================================
                        wbResultExtra.security.workbookPassword = strPW
                        wbResultExtra.security.lockStructure = True
                        filepathResultExtra =  os.path.join(Path(folderOutput),filenameResultExtra)
                        wbResultExtra.save(filepathResultExtra)
                        
            # ========================================================================================
            #    Save LabSamplesYYYY.xlsx workbook
            # ========================================================================================
            wbSample.security.workbookPassword = strPWlab
            wbSample.security.lockStructure = True
            wbSample.save(filepathSampleLocal)
    
        else:
            print('person_db and person sheets absent - Skipping year: '+str(year))
    else:
        print('Could not find: '+filepathSampleLocalOld+' - Skipping year: '+str(year))
       
# =======================================
#    3. Finalization
# =======================================
print('Done!')
        