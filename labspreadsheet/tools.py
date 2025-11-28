#!/usr/bin/env python
import numpy as np
import openpyxl as pyxl
import os
import pandas
from pathlib import Path

from openpyxl.worksheet.datavalidation import DataValidation

###########################
###   Defining styles   ###
###########################
blueFill = pyxl.styles.PatternFill(start_color='CDEEFF',end_color='CDEEFF',fill_type='solid')
greenFill = pyxl.styles.PatternFill(start_color='E5FFDD',end_color='E5FFDD',fill_type='solid')
orangeFill = pyxl.styles.PatternFill(start_color='FFE8DD',end_color='FFE8DD',fill_type='solid')
yellowFill = pyxl.styles.PatternFill(start_color='FFFFCC',end_color='FFFFCC',fill_type='solid')

def AddColumnAnalysis(ws,*,nCol,nRowCorner,nRowHead,nRowContent,iColDateSample,xOnlyDB):
    """AddColumnAnalysis
        
    The purpose of :py:func:`.AddColumnAnalysis` is to update a sheet by adding columns indicating the completion of laboratory analysis.
        
    Parameters
    ----------
    ws : Input sheet
    nCol : Number of columns on the left of the columns that will be added
    nRowCorner : Number of rows used for project-specific information
    nRowHead : Number of rows used as header
    nRowContent : Number of rows available for sample meta-data and analytic results
    iColDateSample : Column number for sample date (first column = 1)
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    # =======================================================
    # Columns: Date, Person
    # =======================================================
    iColDate = nCol+1
    iColPers = nCol+2
    letterDate = pyxl.utils.get_column_letter(iColDate)
    letterPers = pyxl.utils.get_column_letter(iColPers)
    letterDateSample = pyxl.utils.get_column_letter(iColDateSample)
    nCol = nCol+2 
    
    # --------------------------------------
    # Content - Normal formatting
    # --------------------------------------
    ws = ApplyStyles(ws,\
                     c1=iColDate,\
                     c2=iColPers,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead+nRowContent,\
                     addon="")
    
    for iRow in np.arange(nRowCorner+1,nRowCorner+nRowHead+nRowContent):
        cellref = letterDate+str(iRow)
        ws[cellref].style = 'fmtMidLeftDate'
    
    # --------------------------------------
    # Header
    # --------------------------------------
    
    ws.cell(column=iColDate, row=nRowCorner+1, value='Analysis')
    
    ws.cell(column=iColDate, row=nRowCorner+2, value='Date')
    ws.cell(column=iColDate, row=nRowCorner+3, value='[dd.mm.yyyy]')
    ws.cell(column=iColDate, row=nRowCorner+nRowHead, value='Text')
    
    ws.cell(column=iColPers, row=nRowCorner+2, value='Who (A)')
    ws.cell(column=iColPers, row=nRowCorner+nRowHead, value='Text')
    ws = ApplyStyles(ws,\
                     c1=iColDate,\
                     c2=iColPers,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead,\
                     addon="")
    # --------------------------------------
    # Content - Conditional formatting
    # --------------------------------------
    
    iTargetNumber = 2
    inputrange = letterDate+str(nRowCorner+nRowHead+1)+':'+letterPers+str(nRowCorner+nRowHead+1)
    countAll = 'COUNTBLANK('+inputrange+')'
    
    for iCol in range(iTargetNumber):
        letter = pyxl.utils.get_column_letter(iColDate+iCol)
        cellrange = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
        ws.conditional_formatting.add(cellrange,\
                                      pyxl.formatting.rule.FormulaRule(formula=[countAll+'=0'],\
                                                                       fill=greenFill))
        ws.conditional_formatting.add(cellrange,\
                                      pyxl.formatting.rule.FormulaRule(formula=[countAll+'>='+str(iTargetNumber)],\
                                                                       fill=blueFill))
        ws.conditional_formatting.add(cellrange,
                                      \
                                      pyxl.formatting.rule.FormulaRule(formula=[countAll+'<'+str(iTargetNumber)],\
                                                                       fill=orangeFill))
    
    
    # --------------------------------------
    # Content - Validation
    # --------------------------------------
    
    # date validation
    cellsampledate = letterDateSample+str(nRowCorner+nRowHead+1)
    cellref = letterDate+str(nRowCorner+nRowHead+1)
    cellyear = 'B1'
    fDateSample = 'DATE(YEAR('+cellsampledate+'),MONTH('+cellsampledate+'),DAY('+cellsampledate+'))'
    fDateAnalysis = 'DATE(YEAR('+cellref+'),MONTH('+cellref+'),DAY('+cellref+'))'
    
    fDateStart = 'DATE(RIGHT('+cellref+',4),RIGHT(LEFT('+cellref+',5),2),LEFT('+cellref+',2))'
    fDateEnd = 'DATE(RIGHT('+cellref+',4),LEFT(RIGHT('+cellref+',7),2),LEFT(RIGHT('+cellref+',10),2))'
    
    fNotSingleDate = 'ISERROR('+fDateAnalysis+')'
    f3 = 'IF(AND('+fDateStart+'<='+fDateEnd+','+fDateSample+'<='+fDateStart+'),TRUE,FALSE)'
    f2 = 'IF(OR(ISERROR('+fDateStart+'),ISERROR('+fDateEnd+')),FALSE,'+f3+')'
    f1 = 'IF(LEN('+cellref+')=16,'+f2+',FALSE)'
    f0 = 'IF(ISBLANK('+cellsampledate+'),FALSE,IF('+fNotSingleDate+','+f1+',AND(TEXT($B$1,"0")=TEXT(YEAR('+cellref+'),"0"),'+fDateSample+'<='+fDateAnalysis+')))'
    
    dv = DataValidation(type="custom",formula1=f0)
    dv.error ='Your entry should be a date \r\n (a) with format dd.mm.yyyy or dd.mm-dd.mm.yyyy (earliest date first), \r\n (b) with yyyy matching cell B1, and \r\n (c) equal to or after the sample date. \r\n\r\n (!!!) Sample date (column G) must be filled in first'
    dv.errorTitle = 'Invalid date'
    cellrange = letterDate+str(nRowCorner+nRowHead+1)+':'+letterDate+str(nRowCorner+nRowHead+nRowContent)
    dv.add(cellrange)
    ws.add_data_validation(dv)
    
    # person validation
    if xOnlyDB:
        f1str = '=OFFSET(person_db!$B$2,0,0,COUNTIF(person_db!$B$2:$B$1000,">"""),1)'
    else:
        f1str = '=OFFSET(person_drop!$D$2,0,0,COUNTIF(person_drop!$D$2:$D$1000,">"""),1)'
    dv = DataValidation(type="list", formula1=f1str, allow_blank=False)
    dv.error ='Your entry is not in the list of persons'
    dv.errorTitle = 'Invalid person'
    # Optionally set a custom prompt message
    dv.prompt = 'Please select a person from the list'
    dv.promptTitle = 'Select person'
    cellrangeF = letterPers+str(nRowCorner+nRowHead+1)+':'+letterPers+str(nRowCorner+nRowHead+nRowContent)
    dv.add(cellrangeF)
    ws.add_data_validation(dv)
    
    return ws, nCol


def AddColumnComment(ws,*,nCol,nRowCorner,nRowHead,nRowContent,sColumn,xCheck):
    """AddColumnComment
        
    The purpose of :py:func:`.AddColumnComment` is to update a sheet by adding a column where unstructured comments or information is added.
        
    Parameters
    ----------
    ws : Input sheet
    nCol : Number of columns on the left of the columns that will be added
    nRowCorner : Number of rows used for project-specific information
    nRowHead : Number of rows used as header
    nRowContent : Number of rows available for sample meta-data and analytic results
    sColumn : String to place in header of added column
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    nCol = nCol+1
    letter = pyxl.utils.get_column_letter(nCol)
    
    # --------------------------------------
    # Normal formatting: lines 8-1006
    # --------------------------------------
    
    ws = ApplyStyles(ws,\
                     c1=nCol,\
                     c2=nCol,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead+nRowContent,\
                     addon="")
    
    # --------------------------------------
    # Header: lines 4-7
    # --------------------------------------
    ws.cell(column=nCol, row=nRowCorner+1, value=sColumn)
    ws = ApplyStyles(ws,\
                     c1=nCol,\
                     c2=nCol,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead,\
                     addon="")
    
    # --------------------------------------
    # Conditional formatting: lines 8-1006
    # --------------------------------------
    cellrangeApply = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
    if xCheck:
        fill_1 = blueFill
        fill_2 = greenFill
    else:
        fill_1 = yellowFill
        fill_2 = yellowFill
    
    cellrowInput = letter+str(nRowCorner+nRowHead+1)
    ws.conditional_formatting.add(cellrangeApply,\
                                  pyxl.formatting.rule.FormulaRule(formula=['COUNTBLANK('+cellrowInput+')=1'],\
                                                                   fill=fill_1))
    ws.conditional_formatting.add(cellrangeApply,\
                                  pyxl.formatting.rule.FormulaRule(formula=['COUNTBLANK('+cellrowInput+')=0'],\
                                                                   fill=fill_2))
        
    # --------------------------------------
    # Validation: lines 8-1006
    # --------------------------------------
    # None applied
    
    return ws, nCol

def AddColumnConcatID(ws,*,nCol,nRowCorner,nRowHead,nRowContent,strSheet1,strSheet2):
    """AddColumnConcatID
        
    The purpose of :py:func:`.AddColumnConcatID` is to add a column where the first two columns are concatenated. This column facilitates search across Excel sheets. Original implementation was to  automatically fill the columns Room/Area/Setup/Component/Date/Time in the results sheet
        
    Parameters
    ----------
    ws : Input sheet
    nCol : Number of columns on the left of the columns that will be added
    nRowCorner : Number of rows used for project-specific information
    nRowHead : Number of rows used as header
    nRowContent : Number of rows available for sample meta-data and analytic results
    XXX
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    # =======================================================
    # Columns: SampleIDconcat
    # =======================================================
    iColConcatID = nCol+1
    letterConcatID = pyxl.utils.get_column_letter(iColConcatID)
    iColConcatID_X = nCol+2
    letterConcatID_X = pyxl.utils.get_column_letter(iColConcatID_X)
    iColConcatID_Y = nCol+3
    letterConcatID_Y = pyxl.utils.get_column_letter(iColConcatID_Y)
    iColMatchX = nCol+4
    letterMatchX = pyxl.utils.get_column_letter(iColMatchX)
    iColMatchY = nCol+5
    letterMatchY = pyxl.utils.get_column_letter(iColMatchY)
    
    # --------------------------------------
    # Header
    # --------------------------------------
    
    iRow1 = nRowCorner+1
    ws.cell(column=iColConcatID, row=iRow1, value='ConcatID')
    if len(strSheet1)>0:
        ws.cell(column=iColConcatID_X, row=iRow1, value='ConcatID_X')
        ws.cell(column=iColMatchX, row=iRow1, value='MatchX')
    if len(strSheet2)>0:
        ws.cell(column=iColConcatID_Y, row=iRow1, value='ConcatID_Y')
        ws.cell(column=iColMatchY, row=iRow1, value='MatchY')
    
    iRow2 = nRowCorner+2
    iRow3 = nRowCorner+3
    
    # -------------------------------------------------------------------
    # Formatting - always yellow if cells are auto-filled
    # -------------------------------------------------------------------
    
    if len(strSheet1)>0 or len(strSheet2)>0:
        for iCol in np.arange(0,6):
            letter = pyxl.utils.get_column_letter(iCol+3)
            cellrowInput = letter+str(nRowCorner+nRowHead+1)
            cellrangeApply = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
            # This is a roundabout way to fill always with yellow, i.e. by using two rules: could not find a way to make this work with only one rule:
            ws.conditional_formatting.add(cellrangeApply,\
                                  pyxl.formatting.rule.FormulaRule(formula=['COUNTBLANK('+cellrowInput+')=1'],\
                                                                   fill=yellowFill)) 
            ws.conditional_formatting.add(cellrangeApply,\
                                  pyxl.formatting.rule.FormulaRule(formula=['COUNTBLANK('+cellrowInput+')=0'],\
                                                                   fill=yellowFill)) 
    # --------------------------------------
    # Content
    # --------------------------------------
    
    iRowTop = nRowCorner+nRowHead+1
    iRowBottom = nRowCorner+nRowHead+nRowContent+1
    for iRow in np.arange(iRowTop,iRowBottom):
        cellref = letterConcatID+str(iRow)
        fConcat = '=CONCATENATE(A'+str(iRow)+',B'+str(iRow)+')'
        ws[cellref] = fConcat
        if len(strSheet1)>0:
            cellrefX = letterConcatID_X+str(iRow)
            fConcat = '=CONCATENATE(SamplesX!A'+str(iRow)+',SamplesX!B'+str(iRow)+')'
            ws[cellrefX] = fConcat
        if len(strSheet2)>0:
            cellrefY = letterConcatID_Y+str(iRow)
            fConcat = '=CONCATENATE(SamplesY!A'+str(iRow)+',SamplesY!B'+str(iRow)+')'
            ws[cellrefY] = fConcat
        
        celladdX = pyxl.utils.get_column_letter(iColMatchX)+str(iRow)
        celladdY = pyxl.utils.get_column_letter(iColMatchY)+str(iRow)
        
        if len(strSheet1)>0:
            fMatch1 = '=MATCH('+cellref+','+letterConcatID_X+str(1)+':'+letterConcatID_X+str(iRowBottom)+',0)'
            ws[celladdX] = fMatch1
        if len(strSheet2)>0:
            fMatch2 = '=MATCH('+cellref+','+letterConcatID_Y+str(1)+':'+letterConcatID_Y+str(iRowBottom)+',0)'
            ws[celladdY] = fMatch2
        
        if len(strSheet1)>0 or len(strSheet2)>0:
            for iCol in np.arange(0,6):
                strCol = pyxl.utils.get_column_letter(iCol+3)
                fInfo ='=IF(ISNUMBER('+celladdX+'),INDIRECT("SamplesX!'+strCol+'"&'+celladdX+'),IF(ISNUMBER('+celladdY+'),INDIRECT("SamplesY!'+strCol+'"&'+celladdY+'),""))'
                celladdZ = strCol+str(iRow)
                ws[celladdZ] = fInfo
    
    return ws, nCol
    
def AddColumnExpected(ws,*,nCol,nRowCorner,nRowHead,nRowContent,dfVAR):
    """AddColumnExpected
        
    The purpose of :py:func:`.AddColumnExpected` is to update a sheet by adding columns to enter the expected measured values.
        
    Parameters
    ----------
    ws : Input sheet
    nCol : Number of columns on the left of the columns that will be added
    nRowCorner : Number of rows used for project-specific information
    nRowHead : Number of rows used as header
    nRowContent : Number of rows available for sample meta-data and analytic results
    dfVAR : pandas dataframe specifying the meta-data for the measured variables
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    # =======================================================
    # Columns: Room, Area, Unit, Component
    # =======================================================
    
    nVAR = dfVAR.shape[0]
    
    # --------------------------------------
    # Content - Normal formatting
    # --------------------------------------
    
    ws = ApplyStyles(ws,\
                     c1=nCol+1,\
                     c2=nCol+nVAR,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead+nRowContent,\
                     addon="")
    
    # --------------------------------------
    # Header
    # --------------------------------------
    
    ws.cell(column=nCol+1, row=nRowCorner+1, value='Expected value of measured variable')
    for iVAR in np.arange(nVAR):
        ws.cell(column=nCol+1+iVAR, row=nRowCorner+2, value=dfVAR.index[iVAR])
        ws.cell(column=nCol+1+iVAR, row=nRowCorner+3, value=dfVAR['unit'][iVAR])
        ws.cell(column=nCol+1+iVAR, row=nRowCorner+nRowHead, value='Text')
    
    ws = ApplyStyles(ws,\
                     c1=nCol+1,\
                     c2=nCol+nVAR,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead,\
                     addon="")
    
    # --------------------------------------
    # Content - Conditional formatting
    # --------------------------------------
    
    for iVAR in np.arange(nVAR):
        letter = pyxl.utils.get_column_letter(nCol+1+iVAR)
        cellrangeApply = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
        cellrowInput = letter+str(nRowCorner+nRowHead+1)
        ws.conditional_formatting.add(cellrangeApply,\
                                  pyxl.formatting.rule.FormulaRule(formula=['COUNTBLANK('+cellrowInput+')=1'],\
                                                                   fill=blueFill))
        ws.conditional_formatting.add(cellrangeApply,\
                                  pyxl.formatting.rule.FormulaRule(formula=['COUNTBLANK('+cellrowInput+')=0'],\
                                                                   fill=greenFill))
        
    # --------------------------------------
    # Content - Validation
    # --------------------------------------
    # None
    
    nCol = nCol+nVAR
    
    return ws, nCol

def AddColumnID(ws,*,nRowCorner,nRowHead,nRowContent,lstID,strYear,iNumColComplete,xSiteOmitted):
    """AddColumnID
        
    The purpose of :py:func:`.AddColumnID` is to update a sheet by adding columns indicating the project ID and the sample number.
        
    Parameters
    ----------
    ws : Input sheet
    nCol : Number of columns on the left of the columns that will be added
    nRowCorner : Number of rows used for project-specific information
    nRowHead : Number of rows used as header
    nRowContent : Number of rows available for sample meta-data and analytic results
    lstID: List of available project IDs
    strYear: String describing the year for this sheet
    iNumColComplete: The number of columns, starting from the very left of the sheet, that must be filled in for an entry to be considered complete
    xRequired : Boolean indicating whether site columns are required or not
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    # =======================================================
    # Column 1 and 2: Sample ID
    # =======================================================
    iNumID = len(lstID)
    letter1 = pyxl.utils.get_column_letter(1)
    letter2 = pyxl.utils.get_column_letter(2)
    letter7 = pyxl.utils.get_column_letter(7)
    letterEND = pyxl.utils.get_column_letter(iNumColComplete)
    
    # --------------------------------------
    # Normal formatting: lines 8-1006
    # --------------------------------------
    iRowStart = nRowCorner+1 ;
    iRowEnd = nRowCorner+nRowHead+nRowContent ;
    
    ws = ApplyStyles(ws,\
                     c1=1,\
                     c2=2,\
                     r1=iRowStart,\
                     r2=iRowEnd,\
                     addon="")
    
    # --------------------------------------
    # Header: lines 4-7
    # --------------------------------------
    
    ws.cell(column=1, row=nRowCorner+1, value='Sample ID')
    ws.cell(column=1, row=nRowCorner+2, value='Project ID')
    ws.cell(column=1, row=nRowCorner+3, value='[##]')
    ws.cell(column=1, row=nRowCorner+nRowHead, value='Text')
    
    ws.cell(column=2, row=nRowCorner+2, value='Number')
    ws.cell(column=2, row=nRowCorner+3, value='[yysss]')
    ws.cell(column=2, row=nRowCorner+nRowHead, value='Text')
    
    ws = ApplyStyles(ws,\
                     c1=1,\
                     c2=2,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead,\
                     addon="")
    
    # --------------------------------------
    # Conditional formatting: lines 8-1006
    # --------------------------------------
    
    if iNumID==1:
        iTargetNumber = iNumColComplete-1
    else:
        iTargetNumber = iNumColComplete
        
    if xSiteOmitted:
        inputrange1 = letter1+str(nRowCorner+nRowHead+1)+':'+letter2+str(nRowCorner+nRowHead+1)
        inputrange2 = letter7+str(nRowCorner+nRowHead+1)+':'+letterEND+str(nRowCorner+nRowHead+1)
        countAll = 'COUNTBLANK('+inputrange1+')+COUNTBLANK('+inputrange2+')'
        iTargetNumber = iTargetNumber-4
        for iCol in range(3,7):
            letter = pyxl.utils.get_column_letter(iCol)
            ws.column_dimensions[letter].hidden = True
    else:
        inputrange = letter1+str(nRowCorner+nRowHead+1)+':'+letterEND+str(nRowCorner+nRowHead+1)
        countAll = 'COUNTBLANK('+inputrange+')'
    
    for iColumn in range(iNumColComplete):
        letter = pyxl.utils.get_column_letter(iColumn+1)
        cellrange = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
        ws.conditional_formatting.add(cellrange,\
                                      pyxl.formatting.rule.FormulaRule(formula=[countAll+'=0'],\
                                                                       fill=greenFill))
        ws.conditional_formatting.add(cellrange,\
                                      pyxl.formatting.rule.FormulaRule(formula=[countAll+'>='+str(iTargetNumber)],\
                                                                       fill=blueFill))
        ws.conditional_formatting.add(cellrange,\
                                      pyxl.formatting.rule.FormulaRule(formula=[countAll+'<'+str(iTargetNumber)],\
                                                                       fill=orangeFill))
    
    # --------------------------------------
    # Validation: lines 8-1006
    # --------------------------------------
    
    # Column 1
    strID = ','.join(lstID)
    strID = '"'+strID+'"'
    dv = DataValidation(type="list", formula1=strID, allow_blank=True)
    dv.error ='Your entry is not in the list of project IDs for this sheet'
    dv.errorTitle = 'Invalid ID'
    
    cellrange = letter1+str(iRowStart)+':'+letter1+str(iRowEnd)
    dv.add(cellrange)
    ws.add_data_validation(dv)
    
    # Column 2
    N1 = 1000*(int(strYear[2:4])+0)
    N2 = 1000*(int(strYear[2:4])+1)
    cellref = letter2+str(iRowStart)
    f0 = 'ISNUMBER(VALUE('+cellref+'))'
    f1 = 'VALUE('+cellref+')>'+str(N1)
    f2 = 'VALUE('+cellref+')<'+str(N2)
    f3 = 'MOD(VALUE('+cellref+'),1)=0'
    fa = 'AND(AND('+str(f0)+','+str(f3)+'),AND('+str(f1)+','+str(f2)+'))'
    dv = DataValidation(type="custom",formula1=fa,allow_blank=True)
        
    dv.error ='Your sample number should be a whole number between '+str(N1+1)+' and '+str(N2)
    dv.errorTitle = 'Invalid sample number'
    
    cellrange = letter2+str(iRowStart)+':'+letter2+str(iRowEnd)
    dv.add(cellrange)
    ws.add_data_validation(dv)
    
    # Fill in with first option if only one option is available
    '''
    if iNumID==1:
        for iRow in range(nRowContent):
            ws[letter1+str(nRowCorner+nRowHead+iRow+1)].value = lstID[0]
    '''
    
    return ws


def AddColumnMeasurement(ws,*,nCol,nRowCorner,nRowHead,nRowContent,dfVAR):
    """AddColumnMeasurement
        
    The purpose of :py:func:`.AddColumnMeasurement` is to update a sheet by adding columns detailing the laboratory analysis and the resulting value. Per measured variable, four columns are added, in this order: (a) final value, (b) analytic method, (c) applied filter during analysis, and (d) applied dilution during analysis
        
    Parameters
    ----------
    ws : Input sheet
    nCol : Number of columns on the left of the columns that will be added
    nRowCorner : Number of rows used for project-specific information
    nRowHead : Number of rows used as header
    nRowContent : Number of rows available for sample meta-data and analytic results
    dfVAR : pandas dataframe specifying the meta-data for the measured variables
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    iColValue = nCol+1
    iColMethod = nCol+2
    iColFilter = nCol+3
    iColDilu = nCol+4
    
    nVAR = dfVAR.shape[0]
    nCol = nCol+nVAR*4
    
    # --------------------------------------
    # Normal formatting: lines 8-1006
    # --------------------------------------
    
    for iVAR in np.arange(nVAR):
        ws = ApplyStyles(ws,\
                         c1=iColValue+iVAR*4,\
                         c2=iColDilu+iVAR*4,\
                         r1=nRowCorner+1,\
                         r2=nRowCorner+nRowHead+nRowContent,\
                         addon="")
						 
					
    
    # --------------------------------------
    # Header: lines 4-7
    # --------------------------------------
    for iVAR in np.arange(nVAR):
        
        ws.cell(column=iColValue+iVAR*4, row=nRowCorner+1, value=dfVAR.index[iVAR])
        ws.cell(column=iColValue+iVAR*4, row=nRowCorner+2, value='Final value')
        ws.cell(column=iColValue+iVAR*4, row=nRowCorner+3, value=dfVAR['unit'][iVAR])
        ws.cell(column=iColValue+iVAR*4, row=nRowCorner+4, value='Text')
        
        ws.cell(column=iColMethod+iVAR*4, row=nRowCorner+2, value='Method')
        ws.cell(column=iColMethod+iVAR*4, row=nRowCorner+4, value='Dropdown')
        
        ws.cell(column=iColFilter+iVAR*4, row=nRowCorner+2, value='Filtration')
        ws.cell(column=iColFilter+iVAR*4, row=nRowCorner+4, value='Dropdown')
        
        ws.cell(column=iColDilu+iVAR*4, row=nRowCorner+2, value='Dilution')
        ws.cell(column=iColDilu+iVAR*4, row=nRowCorner+4, value='Text')
        
        ws = ApplyStyles(ws,\
                         c1=iColValue+iVAR*4,\
                         c2=iColDilu+iVAR*4,\
                         r1=nRowCorner+1,\
                         r2=nRowCorner+nRowHead,\
                         addon="")
        
        xHidden = True
        for row in ws.iter_cols(min_col=iColValue+iVAR*4, max_col=iColValue+iVAR*4, min_row=nRowCorner+nRowHead+1, max_row=nRowCorner+nRowHead+11, values_only=True):
            for cell in row:
                xHidden = xHidden and (cell==None)
        
        if xHidden:
            letter = pyxl.utils.get_column_letter(iColValue+iVAR*4)
            ws.column_dimensions[letter].hidden = True
            letter = pyxl.utils.get_column_letter(iColMethod+iVAR*4)
            ws.column_dimensions[letter].hidden = True
            letter = pyxl.utils.get_column_letter(iColFilter+iVAR*4)
            ws.column_dimensions[letter].hidden = True
            letter = pyxl.utils.get_column_letter(iColDilu+iVAR*4)
            ws.column_dimensions[letter].hidden = True
        
        
    # --------------------------------------
    # Conditional formatting: lines 8-1006
    # --------------------------------------
    for iVAR in np.arange(nVAR):
    
        lstMethod = dfVAR['method'][iVAR]
        lstFilter = dfVAR['filter'][iVAR]
        
        # conditional formatting
        letter1 = pyxl.utils.get_column_letter(iColValue+iVAR*4)
        letter3 = pyxl.utils.get_column_letter(iColFilter+iVAR*4)
        inputrange = letter1+str(nRowCorner+nRowHead+1)+':'+letter3+str(nRowCorner+nRowHead+1)
        countAll = 'COUNTBLANK('+inputrange+')'
        countValue = 'COUNTBLANK('+letter1+str(nRowCorner+nRowHead+1)+')'
    
        for iColumn in range(4):
            letter = pyxl.utils.get_column_letter(iColValue+iVAR*4+iColumn)
            cellrange = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
            ws.conditional_formatting.add(cellrange,\
                                          pyxl.formatting.rule.FormulaRule(formula=[countAll+'=0'],\
                                                                           fill=greenFill))
            ws.conditional_formatting.add(cellrange,\
                                          pyxl.formatting.rule.FormulaRule(formula=[countValue+'>0'],\
                                                                           fill=blueFill))
            ws.conditional_formatting.add(cellrange,\
                                          pyxl.formatting.rule.FormulaRule(formula=['AND('+countAll+'>0,'+countValue+'=0)'],\
                                                                           fill=orangeFill))
        
        
    # --------------------------------------
    # Validation: lines 8-1006
    # --------------------------------------
    
    for iVAR in np.arange(nVAR):
        
        lstMethod = dfVAR['method'][iVAR]
        lstFilter = dfVAR['filter'][iVAR]
        
        
        # final value validation
        letter = pyxl.utils.get_column_letter(iColValue+iVAR*4)
        cellref = letter+str(nRowCorner+nRowHead+1)
        f1 = 'ISNUMBER(VALUE('+cellref+'))'
        f2 = cellref+'="failed"'
        #f3a = 'EXACT(LEFT('+cellref+',1),"<")'
        f3a = 'OR(EXACT(LEFT('+cellref+',1),"<"),EXACT(LEFT('+cellref+',1),">"))'
        f3b = 'ISNUMBER(VALUE(RIGHT(SUBSTITUTE('+cellref+'," ","_"),LEN('+cellref+')-1)))'
        f3 = '=OR('+f1+',OR('+f2+',AND('+f3a+','+f3b+')))'
        dv = DataValidation(type="custom",formula1=f3)
        dv.error ='Your entry must be (a) a number, (b) "failed", or (c) a string of the form "<42.73" (no spaces)'
        dv.errorTitle = 'Invalid value'
        
        cellrange = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrange)
        ws.add_data_validation(dv)
        
        # method validation
        letter = pyxl.utils.get_column_letter(iColMethod+iVAR*4)
        nMethod = len(lstMethod)
        f1str = ','.join(lstMethod)
        f1str = '"'+f1str+'"'
        dv = DataValidation(type="list", formula1=f1str, allow_blank=False)
        dv.error ='Your entry is not in the list of analytical methods for this variable'
        dv.errorTitle = 'Invalid method'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select an analytical method from the list'
        dv.promptTitle = 'Select method'
        
        cellrangeM = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrangeM)
        ws.add_data_validation(dv)
    
        # filter validation
        letter = pyxl.utils.get_column_letter(iColFilter+iVAR*4)
        nFilter = len(lstFilter)
        f1str = ','.join(lstFilter)
        f1str = '"'+f1str+'"'
        dv = DataValidation(type="list", formula1=f1str, allow_blank=False)
        dv.error ='Your entry is not in the list of filters for this variable'
        dv.errorTitle = 'Invalid filter'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select a filter from the list'
        dv.promptTitle = 'Select filter'
        
        cellrangeF = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrangeF)
        ws.add_data_validation(dv)
    
        # dilution validation
        #letter = pyxl.utils.get_column_letter(iColDilu+iVAR*4)
        #cellref = letter+str(nRowCorner+nRowHead+1)
        #f0 = 'AND(ISNUMBER(VALUE('+cellref+')),MOD('+cellref+',1)=0)'
        #dv = DataValidation(type="custom",formula1=f0)
        #dv = DataValidation(type="whole",operator="greaterThanOrEqual",formula1=1)
        #dv.error ='Your entry should be a whole number equal to or larger than 1'
        #dv.errorTitle = 'Invalid dilution'
    
        #cellrange = letter+str(nRowCorner+nRowHead+1)+':'+letter+str(nRowCorner+nRowHead+nRowContent)
        #dv.add(cellrange)
        #ws.add_data_validation(dv)
        
        for iRow in range(nRowContent):
            
            # Fill in with first option if only one option is available
            if nMethod==1:
                letter = pyxl.utils.get_column_letter(iColMethod+iVAR*4)
                ws[letter+str(nRowCorner+nRowHead+iRow+1)].value = lstMethod[0]
            
            # Fill in default filter option (first option, typically 'None')
            letter = pyxl.utils.get_column_letter(iColFilter+iVAR*4)
            ws[letter+str(nRowCorner+nRowHead+iRow+1)].value = lstFilter[0]
            
            
    return ws, nCol


def AddColumnSample(ws,*,nCol,nRowCorner,nRowHead,nRowContent,lstSampleMethod,lstSampleFilter,xLink,xOnlyDB):
    """AddColumnSample
        
    The purpose of :py:func:`.AddColumnSample` is to update a sheet by adding columns for all meta-data describing a laboratory sample, in this order: (a) date, (b) time, (c) sampling method, (d) applied filter, (e) applied dilution, (f) staff member
        
    Parameters
    ----------
    ws : Input sheet
    nCol : Number of columns on the left of the columns that will be added
    nRowCorner : Number of rows used for project-specific information
    nRowHead : Number of rows used as header
    nRowContent : Number of rows available for sample meta-data and analytic results
    lstSampleMethod : List of available sampling methods
    lstSampleFilter : List of available filters
    xLink : Boolean indicating whether concents are linked (True) or entered manually (False)
    xOnlyDB : Boolean indicating whether person_db (True) or person_drop (False) should be reference for staff members
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    # =======================================================
    # Columns: Date, Time, Filter, Dilution, Person
    # =======================================================
    iColDate = nCol+1
    iColTime = nCol+2
    letterDate = pyxl.utils.get_column_letter(iColDate)
    letterTime = pyxl.utils.get_column_letter(iColTime)
    
    if xLink:
        nCol = nCol+2 # updated number of columns
        # If data is linked - only add columns for Date and Time
    else:
        iColSamp = nCol+3
        iColFilt = nCol+4
        iColDilu = nCol+5
        iColPers = nCol+6
        nCol = nCol+6 # updated number of columns
        letterSamp = pyxl.utils.get_column_letter(iColSamp)
        letterFilt = pyxl.utils.get_column_letter(iColFilt)
        letterDilu = pyxl.utils.get_column_letter(iColDilu)
        letterPers = pyxl.utils.get_column_letter(iColPers)
    
    # --------------------------------------
    # Content - Normal formatting
    # --------------------------------------
    
    ws = ApplyStyles(ws,\
                     c1=iColDate,\
                     c2=nCol,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead+nRowContent,\
                     addon="")
    
    for iRow in np.arange(nRowCorner+1,nRowCorner+nRowHead+nRowContent+1):
        cellref = letterDate+str(iRow)
        ws[cellref].style = 'fmtMidLeftDate'
        cellref = letterTime+str(iRow)
        ws[cellref].style = 'fmtMidCenterTime'
    
    # --------------------------------------
    # Header
    # --------------------------------------
    
    ws.cell(column=iColDate, row=nRowCorner+1, value='Sample')
    
    ws.cell(column=iColDate, row=nRowCorner+2, value='Date')
    ws.cell(column=iColDate, row=nRowCorner+3, value='[dd.mm.yyyy]')
    ws.cell(column=iColDate, row=nRowCorner+nRowHead, value='Text')
    
    ws.cell(column=iColTime, row=nRowCorner+2, value='Time')
    ws.cell(column=iColTime, row=nRowCorner+3, value='[HH:MM:SS]')
    ws.cell(column=iColTime, row=nRowCorner+nRowHead, value='Text')
    
    if xLink:
        # If data is linked - do not add content for Sampling, etc.
        pass
    else:
        ws.cell(column=iColSamp, row=nRowCorner+2, value='Sampling method')
        ws.cell(column=iColSamp, row=nRowCorner+nRowHead, value='Dropdown')
        
        ws.cell(column=iColFilt, row=nRowCorner+2, value='Filter at sampling')
        ws.cell(column=iColFilt, row=nRowCorner+nRowHead, value='Dropdown')
        
        ws.cell(column=iColDilu, row=nRowCorner+2, value='Dilution at sampling')
        ws.cell(column=iColDilu, row=nRowCorner+3, value='[total vol/sample vol]')
        ws.cell(column=iColDilu, row=nRowCorner+nRowHead, value='Text')
        
        ws.cell(column=iColPers, row=nRowCorner+2, value='Who (S)')
        ws.cell(column=iColPers, row=nRowCorner+nRowHead, value='Text')

    ws = ApplyStyles(ws,\
                     c1=iColDate,\
                     c2=nCol,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead,\
                     addon="")
    
    
    # --------------------------------------
    # Content - Conditional formatting
    # --------------------------------------
    # None applied here, see AddColumnID for associated conditional formatting
        
    # --------------------------------------
    # Content - Validation
    # --------------------------------------
    
    if xLink:
        # If data is linked - do not add content here. The function 'AddColumnConcatID' adds the content for the columns Date and Time
        pass
    else:
        # date validation
        cellref = letterDate+str(nRowCorner+nRowHead+1)
        cellyear = 'B1'
        f0 = 'IF(ISERROR(DATE(DAY('+cellref+'),MONTH('+cellref+'),YEAR('+cellref+'))),FALSE,TEXT($B$1,"0")=TEXT(YEAR('+cellref+'),"0"))'
        dv = DataValidation(type="custom",formula1=f0)
        dv.error ='Your entry should be a date (a) with format dd.mm.yyyy and (b) with yyyy matching cell B1'
        dv.errorTitle = 'Invalid date'
        cellrange = letterDate+str(nRowCorner+nRowHead+1)+':'+letterDate+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrange)
        ws.add_data_validation(dv)
        
        # sampling method validation
        nSampleMethod = len(lstSampleMethod)
        f1str = ','.join(lstSampleMethod)
        f1str = '"'+f1str+'"'
        dv = DataValidation(type="list", formula1=f1str, allow_blank=False)
        dv.error ='Your entry is not in the list of sampling methods'
        dv.errorTitle = 'Invalid sampling method'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select a sampling method from the list'
        dv.promptTitle = 'Select sampling method'
        cellrangeF = letterSamp+str(nRowCorner+nRowHead+1)+':'+letterSamp+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrangeF)
        ws.add_data_validation(dv)
        
        # time validation
        cellref = letterTime+str(nRowCorner+nRowHead+1)
        f0 = 'NOT(ISERROR(TIME(HOUR('+cellref+'),MINUTE('+cellref+'),SECOND('+cellref+'))))'
        dv = DataValidation(type="custom",formula1=f0)
        dv.error ='Your entry should be a time with format HH:MM:SS'
        dv.errorTitle = 'Invalid date'
        cellrange = letterTime+str(nRowCorner+nRowHead+1)+':'+letterTime+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrange)
        ws.add_data_validation(dv)
        
        # filter validation
        nFilter = len(lstSampleFilter)
        f1str = ','.join(lstSampleFilter)
        f1str = '"'+f1str+'"'
        dv = DataValidation(type="list", formula1=f1str, allow_blank=False)
        dv.error ='Your entry is not in the list of filters for sample preparation'
        dv.errorTitle = 'Invalid filter'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select a filter from the list'
        dv.promptTitle = 'Select filter'
        cellrangeF = letterFilt+str(nRowCorner+nRowHead+1)+':'+letterFilt+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrangeF)
        ws.add_data_validation(dv)
        
        # dilution validation
        cellref = letterDilu+str(nRowCorner+nRowHead+1)
        f0 = 'AND(ISNUMBER(VALUE('+cellref+')),MOD('+cellref+',1)=0)'
        dv = DataValidation(type="custom",formula1=f0)
        dv.error ='Your entry should be a whole number equal to or larger than 1'
        dv.errorTitle = 'Invalid dilution'
        cellrange = letterDilu+str(nRowCorner+nRowHead+1)+':'+letterDilu+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrange)
        ws.add_data_validation(dv)
        
        # person validation
        if xOnlyDB:
            f1str = '=OFFSET(person_db!$B$2,0,0,COUNTIF(person_db!$B$2:$B$1000,">"""),1)'
        else:
            f1str = '=OFFSET(person_drop!$D$2,0,0,COUNTIF(person_drop!$D$2:$D$1000,">"""),1)'
        dv = DataValidation(type="list", formula1=f1str, allow_blank=False)
        dv.error ='Your entry is not in the list of persons'
        dv.errorTitle = 'Invalid person'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select a person from the list'
        dv.promptTitle = 'Select person'
        cellrangeF = letterPers+str(nRowCorner+nRowHead+1)+':'+letterPers+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrangeF)
        ws.add_data_validation(dv)
        
    return ws, nCol, iColDate

def AddColumnSite(ws,*,nCol,nRowCorner,nRowHead,nRowContent,xLink):
    """AddColumnSite
        
    The purpose of :py:func:`.AddColumnSite` is to update a sheet by adding columns indicating the sampling location (site). A site is specified by four levels of granularity, in this order: (a) room, (b) area, (c) setup, (d) component
        
    Parameters
    ----------
    ws : Input sheet
    nCol : Number of columns on the left of the columns that will be added
    nRowCorner : Number of rows used for project-specific information
    nRowHead : Number of rows used as header
    nRowContent : Number of rows available for sample meta-data and analytic results
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    # =======================================================
    # Columns: Room, Area, Setup, Component
    # =======================================================
    iColRoom  = nCol+1
    iColArea  = nCol+2
    iColSetup = nCol+3
    iColComp  = nCol+4
    nCol      = nCol+4 # updated number of columns
    iOffset   = 50 # number of columns to the right to place intermediate computations in 
    
    letterRoom = pyxl.utils.get_column_letter(iColRoom)
    letterRoomOff = pyxl.utils.get_column_letter(iColRoom+iOffset)
    letterArea = pyxl.utils.get_column_letter(iColArea)
    letterAreaOff = pyxl.utils.get_column_letter(iColArea+iOffset)
    letterSetup= pyxl.utils.get_column_letter(iColSetup)
    letterSetupOff= pyxl.utils.get_column_letter(iColSetup+iOffset)
    letterComp = pyxl.utils.get_column_letter(iColComp)
    
    # --------------------------------------
    # Normal formatting: lines 8-1006
    # --------------------------------------
    
    ws = ApplyStyles(ws,\
                         c1=iColRoom,\
                         c2=iColComp,\
                         r1=nRowCorner+1,\
                         r2=nRowCorner+nRowHead+nRowContent,\
                         addon="")
    
    # --------------------------------------
    # Header:
    # --------------------------------------
    
    # columns for data entry:
    ws.cell(column=iColRoom, row=nRowCorner+1, value='Site')
    
    ws.cell(column=iColRoom, row=nRowCorner+2, value='Room')
    ws.cell(column=iColArea, row=nRowCorner+2, value='Area')
    ws.cell(column=iColSetup, row=nRowCorner+2, value='Setup')
    ws.cell(column=iColComp, row=nRowCorner+2, value='Component')
    
    ws.cell(column=iColRoom, row=nRowCorner+nRowHead, value='Dropdown')
    ws.cell(column=iColArea, row=nRowCorner+nRowHead, value='Dropdown')
    ws.cell(column=iColSetup, row=nRowCorner+nRowHead, value='Dropdown')
    ws.cell(column=iColComp, row=nRowCorner+nRowHead, value='Dropdown')
    
    # columns for intermediate computations:
    if xLink:
        pass
    else:
        ws.cell(column=iColArea+iOffset-1, row=nRowCorner+1, value='Number of options')
    
        ws.cell(column=iColArea+iOffset-1, row=nRowCorner+2, value='Area')
        ws.cell(column=iColSetup+iOffset-1, row=nRowCorner+2, value='Setup')
        ws.cell(column=iColComp+iOffset-1, row=nRowCorner+2, value='Component')
    
    ws = ApplyStyles(ws,\
                     c1=iColRoom,\
                     c2=iColComp,\
                     r1=nRowCorner+1,\
                     r2=nRowCorner+nRowHead,\
                     addon="")

    # --------------------------------------
    # Conditional formatting: lines 8-1006
    # --------------------------------------
    # None applied here, see 'AddColumnID' for conditional formatting associated with these columns
    
    # --------------------------------------
    # Validation: lines 8-1006
    # --------------------------------------
    
    if xLink:
        # If data is linked - do not add content here. The function 'AddColumnConcatID' adds the content for the columns Date and Time
        pass
    else:
        # room validation
        f1str = 'OFFSET(Area!$A$1,0,0,1,COUNTA(Area!$1:$1))'
        dv = DataValidation(type="list", formula1=f1str, allow_blank=False)
        dv.error ='Your entry is not in the list of rooms'
        dv.errorTitle = 'Invalid room'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select a room from the list'
        dv.promptTitle = 'Select room'
        cellrangeF = letterRoom+str(nRowCorner+nRowHead+1)+':'+letterRoom+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrangeF)
        ws.add_data_validation(dv)
    
        # area validation - count available options
        for iRow in np.arange(nRowCorner+nRowHead+1,nRowCorner+nRowHead+nRowContent+1):
            strSourceRoom = letterRoom+str(iRow)
            strTarget = letterRoomOff+str(iRow)
            fColumnRef = 'SUBSTITUTE(ADDRESS(1,MATCH('+strSourceRoom+',Area!$1:$1,0),4),"1","")'
            fCount = '=IF(COUNTBLANK('+strSourceRoom+')=0,COUNTA(INDIRECT(CONCATENATE("Area!",'+fColumnRef+',"2:",'+fColumnRef+',"99"))),0)'
            ws[strTarget] = fCount
    
        # area validation - actual validation
        iRow = nRowCorner+nRowHead+1
        strSourceRoom = letterRoom+str(iRow)
        strTarget = letterRoomOff+str(iRow)
        fColumnRef = 'SUBSTITUTE(ADDRESS(1,MATCH('+strSourceRoom+',Area!$1:$1,0),4),"1","")'
        f0 = '=INDIRECT(CONCATENATE("Area!",'+fColumnRef+',"2:",'+fColumnRef+',TEXT(1+'+strTarget+',"0")))'
        dv = DataValidation(type="list", formula1=f0, allow_blank=False)
        dv.error ='Your entry is not in the list of areas for the selected room'
        dv.errorTitle = 'Invalid area'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select an area from the list'
        dv.promptTitle = 'Select area'
        
        cellrangeF = letterArea+str(nRowCorner+nRowHead+1)+':'+letterArea+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrangeF)
        ws.add_data_validation(dv)
    
        # setup validation - count available options
        for iRow in np.arange(nRowCorner+nRowHead+1,nRowCorner+nRowHead+nRowContent+1):
            strSourceRoom = letterRoom+str(iRow)
            strSourceArea = letterArea+str(iRow)
            strTarget = letterAreaOff+str(iRow)
            fColumnRef = 'SUBSTITUTE(ADDRESS(1,MATCH(CONCATENATE('+strSourceRoom+',":",'+strSourceArea+'),Setup!$1:$1,0),4),"1","")'
            fCount = '=IF(COUNTBLANK('+strSourceArea+')=0,COUNTA(INDIRECT(CONCATENATE("Setup!",'+fColumnRef+',"2:",'+fColumnRef+',"99"))),0)'
            ws[strTarget] = fCount
        
        # setup validation - actual validation
        iRow = nRowCorner+nRowHead+1
        strSourceRoom = letterRoom+str(iRow)
        strSourceArea = letterArea+str(iRow)
        strTarget = letterAreaOff+str(iRow)
        fColumnRef = 'SUBSTITUTE(ADDRESS(1,MATCH(CONCATENATE('+strSourceRoom+',":",'+strSourceArea+'),Setup!$1:$1,0),4),"1","")'
        f0 = '=INDIRECT(CONCATENATE("Setup!",'+fColumnRef+',"2:",'+fColumnRef+',TEXT(1+'+strTarget+',"0")))'
        dv = DataValidation(type="list", formula1=f0, allow_blank=False)
        dv.error ='Your entry is not in the list of areas for the selected room'
        dv.errorTitle = 'Invalid area'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select an area from the list'
        dv.promptTitle = 'Select area'
    
        cellrangeF = letterSetup+str(nRowCorner+nRowHead+1)+':'+letterSetup+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrangeF)
        ws.add_data_validation(dv)
    
        # component validation - count available options
        for iRow in np.arange(nRowCorner+nRowHead+1,nRowCorner+nRowHead+nRowContent+1):
            strSourceRoom = letterRoom+str(iRow)
            strSourceArea = letterArea+str(iRow)
            strSourceSetup = letterSetup+str(iRow)
            strTarget = letterSetupOff+str(iRow)
            fColumnRef = 'SUBSTITUTE(ADDRESS(1,MATCH(CONCATENATE('+strSourceRoom+',":",'+strSourceArea+',":",'+strSourceSetup+'),Comp!$1:$1,0),4),"1","")'
            fCount = '=IF(COUNTBLANK('+strSourceSetup+')=0,COUNTA(INDIRECT(CONCATENATE("Comp!",'+fColumnRef+',"2:",'+fColumnRef+',"99"))),0)'
            ws[strTarget] = fCount
        
        # component validation - actual validation
        iRow = nRowCorner+nRowHead+1
        strSourceRoom = letterRoom+str(iRow)
        strSourceArea = letterArea+str(iRow)
        strSourceSetup = letterSetup+str(iRow)
        strTarget = letterSetupOff+str(iRow)
        fColumnRef = 'SUBSTITUTE(ADDRESS(1,MATCH(CONCATENATE('+strSourceRoom+',":",'+strSourceArea+',":",'+strSourceSetup+'),Comp!$1:$1,0),4),"1","")'
        f0 = '=INDIRECT(CONCATENATE("Comp!",'+fColumnRef+',"2:",'+fColumnRef+',TEXT(1+'+strTarget+',"0")))'
        #print(f0)
        dv = DataValidation(type="list", formula1=f0, allow_blank=False)
        dv.error ='Your entry is not in the list of setups for the selected room'
        dv.errorTitle = 'Invalid setup'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select a setup from the list'
        dv.promptTitle = 'Select setup'
        
        cellrangeF = letterComp+str(nRowCorner+nRowHead+1)+':'+letterComp+str(nRowCorner+nRowHead+nRowContent)
        dv.add(cellrangeF)
        ws.add_data_validation(dv)
        
    return ws, nCol

def AddCorner(ws,*,ProjectInfo,xSample,filename,filenameAdd,folder):
    """AddCorner
        
    The purpose of :py:func:`.AddCorner` is to update a sheet by adding a corner section which includes the project meta-data and links to related files.
        
    Parameters
    ----------
    ws : Input sheet (assumed blank)
    ProjectInfo : dictionary specifying project meta-data
    xSample : boolean indicating whether the input sheet is a sample sheet (True) or a result sheet (False)
    filename : sample/result file for which a hyperlink is provided
    filenameAdd : second sample/result file for which a hyperlink is provided
    folder : folder where file "filename" is located
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    strProjectName = ProjectInfo.Project 
    strEmail = ProjectInfo.Email
    strYear = ProjectInfo.Year
    strFolder = ProjectInfo.Folder
    
    if xSample:
        topic = '[Lab samples]'
        linkedtype = 'Results: '
    else:
        topic = '[Lab results]'
        linkedtype = 'Samples: '
        
    link0  = '=HYPERLINK("mailto:'+strEmail+'?subject='+topic+'", "'+strEmail+'")'
    if len(filename)>0:
        if len(folder)==0:
            folder =strFolder
        link1= '=HYPERLINK("'+folder+'\\'+filename+'#'+strProjectName+'!A1","'+filename+'")'
    else:
        link1 = ""
    if len(filenameAdd)>0:
        link2= '=HYPERLINK("'+strFolder+'\\'+filenameAdd+'#extra_'+strProjectName+'!A1","'+filenameAdd+'")'
    else:
        link2 = ""
        
    ws.cell(column=1, row=1, value=strProjectName)
    ws.cell(column=1, row=3, value=linkedtype)
    ws.cell(column=1, row=2, value='Responsible: ')
    
    ws.cell(column=2, row=1, value=strYear)
    ws.cell(column=2, row=2, value=link0)
    ws.cell(column=2, row=3, value=link1)
    ws.cell(column=2, row=4, value=link2)
    
    nRowCorner = 4 
    
    ws = ApplyStyles(ws,\
                         c1=1,\
                         c2=2,\
                         r1=1,\
                         r2=nRowCorner,\
                         addon='Corner')
    
    
    return ws, nRowCorner

def AddStyles(wb):
    """AddStyles
        
    The purpose of :py:func:`.AddStyles` is to update a workbook by adding format styles
        
    Parameters
    ----------
    wb : Input workbook
    
    Returns
    -------
    wb : Updated workbook
    
    """
    
    bdTL = pyxl.styles.Border(top=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thin",color='FF000000'))
    bdML = pyxl.styles.Border(top=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thin",color='FF000000'))
    bdBL = pyxl.styles.Border(top=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thin",color='FF000000'))
    bdTC = pyxl.styles.Border(top=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thin",color='FF000000'))
    bdMC = pyxl.styles.Border(top=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thin",color='FF000000'))
    bdBC = pyxl.styles.Border(top=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thin",color='FF000000'))
    bdTR = pyxl.styles.Border(top=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thick",color='FF000000'))
    bdMR = pyxl.styles.Border(top=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thick",color='FF000000'))
    bdBR = pyxl.styles.Border(top=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thin",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thick",color='FF000000'))
    bdTLC= pyxl.styles.Border(top=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="none",color='FF000000'))
    bdMLC= pyxl.styles.Border(top=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="none",color='FF000000'))
    bdBLC= pyxl.styles.Border(top=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="none",color='FF000000'))
    bdTRC= pyxl.styles.Border(top=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thick",color='FF000000'))
    bdMRC= pyxl.styles.Border(top=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thick",color='FF000000'))
    bdBRC= pyxl.styles.Border(top=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          left=pyxl.styles.Side(border_style="none",color='FF000000'),\
                          bottom=pyxl.styles.Side(border_style="thick",color='FF000000'),\
                          right=pyxl.styles.Side(border_style="thick",color='FF000000'))
    
    alignTopLeft = pyxl.styles.Alignment(horizontal='left',vertical='top')
    alignContent = pyxl.styles.Alignment(horizontal='center',vertical='top')
    
    formatTL = pyxl.styles.NamedStyle(name="fmtTopLeft")
    formatTL.font = pyxl.styles.Font(bold=True, size=11)
    formatTL.number_format = '@'
    formatTL.alignment = alignTopLeft
    formatTL.border = bdTL
    
    formatML = pyxl.styles.NamedStyle(name="fmtMidLeft")
    formatML.font = pyxl.styles.Font(bold=False, size=11)
    formatML.number_format = '@'
    formatML.border = bdML
    formatML.alignment = alignContent
    
    formatMLdate = pyxl.styles.NamedStyle(name="fmtMidLeftDate")
    formatMLdate.font = pyxl.styles.Font(bold=False, size=11)
    formatMLdate.number_format = 'DD.MM.YYYY'
    formatMLdate.border = bdML
    formatMLdate.alignment = alignContent
    
    formatBL = pyxl.styles.NamedStyle(name="fmtBottomLeft")
    formatBL.font = pyxl.styles.Font(bold=False, size=11)
    formatBL.number_format = '@'
    formatBL.border = bdBL
    formatBL.alignment = alignContent
    
    formatTC = pyxl.styles.NamedStyle(name="fmtTopCenter")
    formatTC.font = pyxl.styles.Font(bold=False, size=11)
    formatTC.number_format = '@'
    formatTC.border = bdTC
    formatTC.alignment = alignContent
    
    formatMC = pyxl.styles.NamedStyle(name="fmtMidCenter")
    formatMC.font = pyxl.styles.Font(bold=False, size=11)
    formatMC.number_format = '@'
    formatMC.border = bdMC
    formatMC.alignment = alignContent
    
    formatMCtime = pyxl.styles.NamedStyle(name="fmtMidCenterTime")
    formatMCtime.font = pyxl.styles.Font(bold=False, size=11)
    formatMCtime.number_format = 'HH:MM:SS'
    formatMCtime.border = bdMC
    formatMCtime.alignment = alignContent
    
    formatBC = pyxl.styles.NamedStyle(name="fmtBottomCenter")
    formatBC.font = pyxl.styles.Font(bold=False, size=11)
    formatBC.number_format = '@'
    formatBC.border = bdBC
    formatBC.alignment = alignContent
    
    formatTR = pyxl.styles.NamedStyle(name="fmtTopRight")
    formatTR.font = pyxl.styles.Font(bold=False, size=11)
    formatTR.number_format = '@'
    formatTR.border = bdTR
    formatTR.alignment = alignContent
    
    formatMR = pyxl.styles.NamedStyle(name="fmtMidRight")
    formatMR.font = pyxl.styles.Font(bold=False, size=11)
    formatMR.number_format = '@'
    formatMR.border = bdMR
    formatMR.alignment = alignContent
    
    formatBR = pyxl.styles.NamedStyle(name="fmtBottomRight")
    formatBR.font = pyxl.styles.Font(bold=False, size=11)
    formatBR.number_format = '@'
    formatBR.border = bdBR
    formatBR.alignment = alignContent
    
    formatTLC = pyxl.styles.NamedStyle(name="fmtTopLeftCorner")
    formatTLC.font = pyxl.styles.Font(bold=False, size=11,color='FF000099')
    formatTLC.number_format = '@'
    formatTLC.border = bdTLC
    formatTLC.alignment = alignContent
    
    formatMLC = pyxl.styles.NamedStyle(name="fmtMidLeftCorner")
    formatMLC.font = pyxl.styles.Font(bold=False, size=11,color='FF000099')
    formatMLC.number_format = '@'
    formatMLC.border = bdMLC
    formatMLC.alignment = alignContent
    
    formatBLC = pyxl.styles.NamedStyle(name="fmtBottomLeftCorner")
    formatBLC.font = pyxl.styles.Font(bold=False, size=11,color='FF000099')
    formatBLC.number_format = '@'
    formatBLC.border = bdBLC
    formatBLC.alignment = alignContent

    formatTRC = pyxl.styles.NamedStyle(name="fmtTopRightCorner")
    formatTRC.font = pyxl.styles.Font(bold=False, size=11,color='FF000099')
    formatTRC.number_format = '@'
    formatTRC.border = bdTRC
    formatTRC.alignment = alignContent
    
    formatMRC = pyxl.styles.NamedStyle(name="fmtMidRightCorner")
    formatMRC.font = pyxl.styles.Font(bold=False, size=11,color='FF000099',underline="single")
    formatMRC.number_format = '@'
    formatMRC.border = bdMRC
    formatMRC.alignment = alignContent
    
    formatBRC = pyxl.styles.NamedStyle(name="fmtBottomRightCorner")
    formatBRC.font = pyxl.styles.Font(bold=False, size=11,color='FF000099',underline="single")
    formatBRC.number_format = '@'
    formatBRC.border = bdBRC
    formatBRC.alignment = alignContent
   
    wb.add_named_style(formatTL)
    wb.add_named_style(formatML)
    wb.add_named_style(formatMLdate)
    wb.add_named_style(formatBL)
    wb.add_named_style(formatTC)
    wb.add_named_style(formatMC)
    wb.add_named_style(formatMCtime)
    wb.add_named_style(formatBC)
    wb.add_named_style(formatTR)
    wb.add_named_style(formatMR)
    wb.add_named_style(formatBR)
    wb.add_named_style(formatTLC)
    wb.add_named_style(formatMLC)
    wb.add_named_style(formatBLC)
    wb.add_named_style(formatTRC)
    wb.add_named_style(formatMRC)
    wb.add_named_style(formatBRC)
    
    return wb

def AddTableSite(wb,*,dfSITE):
    """AddTableSite
        
    The purpose of :py:func:`.AddTableSite` is to update a workbook by adding sheets specifying the available site specifications. These are used to program the dependent drop-down menus for site specification without resorting to visual basic (by design choice) or array formulas (as this requires manual entry of Ctrl-Alt-Shift).
        
    Parameters
    ----------
    wb : Input workbook
    dfSITE : pandas dataframe  specifying the available sampling sites
    
    Returns
    -------
    wb : Updated workbook
    
    """
    
    for iLevel in range(dfSITE.shape[1]-1):
        wb.create_sheet(dfSITE.columns[iLevel+1])
        ws = wb[dfSITE.columns[iLevel+1]]
    
        Grps = dfSITE.groupby(list(dfSITE.columns[0:iLevel+1])) #.groups
        for iCol, Grp in enumerate(Grps):
            if iLevel==0:
                N =1
            else: 
                N=len(Grp[0])
        
            strHead = ''
            for iRow in range(N):
                if iLevel==0:
                    v = Grp[0]
                else:
                    v = Grp[0][iRow]
                strHead = strHead+':'+v
            
            strHead = strHead[1:]
            letter = pyxl.utils.get_column_letter(iCol+1)
            ws[letter+str(1)] = strHead
            
            lstChoices = list(Grp[1].iloc[:,iLevel+1].drop_duplicates())
            for iRow, v in enumerate(lstChoices):
                ws[letter+str(iRow+2)] = v
        ws.sheet_state = 'hidden'
        ws.protection.sheet = True
        ws.protection.formatColumns = False
        ws.protection.set_password('openpyxl')
        
    return wb

def ApplyStyles(ws,*,c1,c2,r1,r2,addon):
    """ApplyStyles
        
    The purpose of :py:func:`.ApplyStyles` is to update a sheet by applying format styles to a range of cells
        
    Parameters
    ----------
    ws : Input sheet
    c1 : First column
    c2 : Last column
    r1 : First row
    r2 : Last row
    addon : String to append to format style specification (e.g., `date` and `time`)
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    for iCol in np.arange(c1,c2+1):
        if iCol==c1:
            sCol = 'Left'
        elif (iCol==c2): 
            sCol = 'Right'
        else:
            sCol = 'Center'
        
        for iRow in np.arange(r1,r2+1):
            if iRow==r1:
                sRow = 'Top'
            elif (iRow==r2):
                sRow = 'Bottom'
            else:
                sRow = 'Mid'
            
            
            letter = pyxl.utils.get_column_letter(iCol)
            cellref = letter+str(iRow)
            fmt = 'fmt'+sRow+sCol+addon
            ws[cellref].style = fmt
        
    return ws

def ConstructLinks(ws,*,strFolder,strFilename,strSheet,nRow,nCol):
    """ConstructLinks
        
    The purpose of :py:func:`.ConstructLinks` is to update a sheet by adding links to entries in a sheet with name `person` in another file
        
    Parameters
    ----------
    ws : Input sheet (assumed blank)
    strFolder : folder where target file is located 
    strFilename : filename of file with sheet `person` managed by laboratory staff
    strSheet : name of sheet that needs to be linked
    nRow : number of rows that need to be copied (starting from top)
    nCol : number of columns that need to be copied (starting from left)
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    ws.sheet_state = 'hidden'
    for iRow in range(nRow):
        for iCol in range(nCol):
            strCell = pyxl.utils.get_column_letter(iCol+1)+str(iRow+1)
            link = "'"+strFolder+"\["+strFilename+"]"+strSheet+"'!"+strCell
            ws[strCell]= '=IF(ISBLANK('+link+')," ",'+link+')'
    
    return ws
    
def ConstructPersonDrop(ws):
    """ConstructPersonDrop
        
    The purpose of :py:func:`.ConstructPersonDrop` is to create a sheet which lists a combined set of entries in two other sheets (specifically: column D in new sheet includes all entries in column B of the sheets 'person' and 'person_db')
        
    Parameters
    ----------
    ws : Input sheet (assumed blank)
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    # A1 cell
    fCOUNT = '=COUNTIF(person_db!$B$2:$B$1000,">""")'
    ws['A1'] = fCOUNT
    
    # B+C cells
    ws['B1'] = 0
    for iRow in np.arange(1,1000):
        cellrefB = 'B'+str(iRow+1)
        ws[cellrefB] = iRow
        fSELECT= '='+str(cellrefB)+'<=A1'
        cellrefC = 'C'+str(iRow+1)
        ws[cellrefC] = fSELECT
        cellrefD = 'D'+str(iRow+1)
        fALIAS='=IF(C'+str(iRow+1)+',INDIRECT("person_db!B"&('+cellrefB+'+1)),INDIRECT("person!B"&('+cellrefB+'-$A$1+1)))'
        ws[cellrefD] = fALIAS
        
        Protect(ws,nRowCorner=0,nRowHead=0,nCol=0)
        ws.sheet_state = 'hidden'
    
    return ws
    
def ConstructResultSheet(wsRes,*,ProjectInfo,folder,filename,filenameExtra,dfVAR,xLab):    
    """ConstructResultSheet
        
    The purpose of :py:func:`.ConstructResultSheet` is to update a sheet by specifying all columns necessary to describe all analytical results and the methods used to obtain them
        
    Parameters
    ----------
    wsRes : Input sheet (assumed blank)
    ProjectInfo : dictionary specifying project meta-data
    folder : folder name for file 'filename' (string)
    filename : sample filename
    filenameExtra : second sample filename
    dfVAR : pandas dataframe specifying the meta-data for the measured variables
    xLab : boolean indicating whether the specified sheet is for samples analyzed by the lab staff (True) or project staff (False)
    
    Returns
    -------
    wsRes : Updated sheet
    
    """
    
    nRowContent = 999 # Hard-coded number of number of samples
    nRowHead = 4 # Hard-coded number of header lines (excluding top-left corner cells)
    strYear = str(ProjectInfo.Year)
    if xLab:
        lstSampleID = ProjectInfo.SampleID_Lab.split(', ')
    else:
        lstSampleID = ProjectInfo.SampleID_Project.split(', ')
    
    # ==============================
    # Top-left corner
    # ==============================
    wsRes,nRowCorner = AddCorner(wsRes,\
                              ProjectInfo=ProjectInfo,\
                              xSample=False,\
                              filename=filename,\
                              filenameAdd=filenameExtra,\
                              folder=folder)

    nColRes=2 # hard-coded number of columns on the left of the next columns.
    wsRes = AddColumnID(wsRes,\
                     nRowCorner=nRowCorner,\
                     nRowHead=nRowHead,\
                     nRowContent=nRowContent,\
                     lstID=lstSampleID,\
                     strYear=strYear,\
                     iNumColComplete=nColRes,\
                     xSiteOmitted=False)
    wsRes,nColRes = AddColumnSite(wsRes,\
                            nCol=nColRes,\
                            nRowCorner=nRowCorner,\
                            nRowHead=nRowHead,\
                            nRowContent=nRowContent,xLink=True)
    wsRes,nColRes,iColDate = AddColumnSample(wsRes,\
                                             nCol=nColRes,\
                                             nRowCorner=nRowCorner,\
                                             nRowHead=nRowHead,\
                                             nRowContent=nRowContent,\
                                             lstSampleMethod=[],\
                                             lstSampleFilter=[],\
                                             xLink=True,\
                                             xOnlyDB=False)
    xSiteOmitted = (ProjectInfo.SiteIncluded=='False')
    wsRes,nColRes = AddColumnComment(wsRes,\
                               nCol=nColRes,\
                               nRowCorner=nRowCorner,\
                               nRowHead=nRowHead,\
                               nRowContent=nRowContent,\
                               sColumn='SampleInfo',xCheck=False)
    wsRes,nColRes = AddColumnComment(wsRes,\
                               nCol=nColRes,\
                               nRowCorner=nRowCorner,\
                               nRowHead=nRowHead,\
                               nRowContent=nRowContent,\
                               sColumn='Comment',xCheck=True)
    wsRes,nColRes = AddColumnComment(wsRes,\
                               nCol=nColRes,\
                               nRowCorner=nRowCorner,\
                               nRowHead=nRowHead,\
                               nRowContent=nRowContent,\
                               sColumn='Communication Lab',xCheck=False)
    wsRes,nColRes = AddColumnMeasurement(wsRes,\
                              nCol=nColRes,\
                              nRowCorner=nRowCorner,\
                              nRowHead=nRowHead,\
                              nRowContent=nRowContent,\
                              dfVAR=dfVAR)
    if len(filename)>0:
        strSheet1="'"+folder+'\\['+filename+']'+ProjectInfo.Project+"'"
    else:
        strSheet1=""
        
    if len(filenameExtra)>0:
        strSheet2="'"+ProjectInfo.Folder+'\\['+filenameExtra+']extra_'+ProjectInfo.Project+"'"
    else:
        strSheet2=""
        
    wsSam,nColRes = AddColumnConcatID(wsRes,\
                                      nCol=nColRes+1,\
                                      nRowCorner=nRowCorner,\
                                      nRowHead=nRowHead,\
                                      nRowContent=nRowContent,\
                                      strSheet1=strSheet1,
                                      strSheet2=strSheet2)
    
    wsRes = Protect(wsRes,nRowCorner=nRowCorner,nRowHead=nRowHead,nCol=nColRes)
    
    # Protect columns with linked meta-data:
    nRow = wsRes.max_row
    for iRow in range(nRow-nRowCorner-nRowHead):
        for iCol in range(6):
            letter = pyxl.utils.get_column_letter(iCol+3)
            cell = wsRes[letter+str(nRowCorner+nRowHead+iRow+1)]
            cell.protection = pyxl.styles.Protection(locked=True)
            
    return wsRes

def ConstructSampleSheet(wsSam,*,ProjectInfo,filename,filenameExtra,lstSampleMethod,lstSampleFilter,dfVAR,xLab):
    """ConstructSampleSheet
        
    The purpose of :py:func:`.ConstructSampleSheet` is to update a sheet by specifying all columns necessary to describe all meta-data describing a laboratory sample
        
    Parameters
    ----------
    wsSam : Input sheet (assumed blank)
    ProjectInfo : dictionary specifying project meta-data
    filename : result filename
    filenameExtra : second result filename
    lstSampleMethod : list of sampling methods
    lstSampleFilter :  list of filters available at sampling
    dfVAR : pandas dataframe specifying the meta-data for the measured variables
    xLab : boolean indicating whether the specified sheet is for samples analyzed by the lab staff (True) or project staff (False)
    
    Returns
    -------
    wsSam : Updated sheet
    
    """
    
    nRowContent = 999 # Hard-coded number of number of samples
    nRowHead = 4 # Hard-coded number of header lines (excluding top-left corner cells)
    strYear = str(ProjectInfo.Year)
    if xLab:
        lstSampleID = ProjectInfo.SampleID_Lab.split(', ')
    else:
        lstSampleID = ProjectInfo.SampleID_Project.split(', ')
    
    wsSam,nRowCorner = AddCorner(wsSam,\
                              ProjectInfo=ProjectInfo,\
                              xSample=True,\
                              filename=filename,\
                              filenameAdd=filenameExtra,\
                              folder="")
    
    nColSam=2 # hard-coded number of columns on the left of the next columns.
    wsSam,nColSam = AddColumnSite(wsSam,\
                            nCol=nColSam,\
                            nRowCorner=nRowCorner,\
                            nRowHead=nRowHead,\
                            nRowContent=nRowContent,xLink=False)
    wsSam,nColSam,iColDate = AddColumnSample(wsSam,\
                                             nCol=nColSam,\
                                             nRowCorner=nRowCorner,\
                                             nRowHead=nRowHead,\
                                             nRowContent=nRowContent,\
                                             lstSampleMethod=lstSampleMethod,\
                                             lstSampleFilter=lstSampleFilter,\
                                             xLink=False,\
                                             xOnlyDB=(ProjectInfo.OnlyDB=='True'))
    xSiteOmitted = (ProjectInfo.SiteIncluded=='False')
    wsSam = AddColumnID(wsSam,\
                     nRowCorner=nRowCorner,\
                     nRowHead=nRowHead,\
                     nRowContent=nRowContent,\
                     lstID=lstSampleID,\
                     strYear=strYear,\
                     iNumColComplete=nColSam,\
                     xSiteOmitted=xSiteOmitted )
    
    wsSam,nColSam = AddColumnComment(wsSam,\
                               nCol=nColSam,\
                               nRowCorner=nRowCorner,\
                               nRowHead=nRowHead,\
                               nRowContent=nRowContent,\
                               sColumn='Comment',xCheck=True)
    wsSam,nColSam = AddColumnComment(wsSam,\
                               nCol=nColSam,\
                               nRowCorner=nRowCorner,\
                               nRowHead=nRowHead,\
                               nRowContent=nRowContent,\
                               sColumn='Communication Lab',xCheck=False)
    wsSam,nColSam = AddColumnAnalysis(wsSam,\
                                      nCol=nColSam,\
                                      nRowCorner=nRowCorner,\
                                      nRowHead=nRowHead,\
                                      nRowContent=nRowContent,\
                                      iColDateSample=iColDate,\
                                      xOnlyDB=(ProjectInfo.OnlyDB=='True'))
    wsSam,nColSam = AddColumnExpected(wsSam,\
                                nCol=nColSam,\
                                nRowCorner=nRowCorner,\
                                nRowHead=nRowHead,\
                                nRowContent=nRowContent,\
                                dfVAR=dfVAR)
    wsSam,nColSam = AddColumnConcatID(wsSam,\
                                nCol=nColSam+1,\
                                nRowCorner=nRowCorner,\
                                nRowHead=nRowHead,\
                                nRowContent=nRowContent,strSheet1="",strSheet2="")
    
    #ws = AdjustWidth(ws)
    wsSam = Protect(wsSam,nRowCorner=nRowCorner,nRowHead=nRowHead,nCol=nColSam)
    
    return wsSam


def CopyData(ws,*,folderInput,filename,strSheetName,row1,xCopyStyle):
    """CopyData
        
    The purpose of :py:func:`.CopyData` is to make copies of existing data into a new sheet, including some style elements if wanted
        
    Parameters
    ----------
    ws : Input sheet (assumed blank)
    folderInput : folder that contains file with source
    filename : name of file with source data
    strSheetName : sheet name of sheet with source data
    row1 : first row that should be copied (count starts at 1)
    xCopyStyle : boolean indicating whether style elements should be copied (font style, cell overline and underline)
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    filenameOld =  os.path.join(Path(folderInput),filename)
    if os.path.isfile(filenameOld):
        #print(filenameOld)
        wbOld = pyxl.load_workbook(filenameOld) #Add file name
        if strSheetName in wbOld.sheetnames:
            wsOld = wbOld[strSheetName] #Add Sheet name
            for iCol in range(200):
                letter = pyxl.utils.get_column_letter(iCol+1)
                for iRow in range(row1,1007):
                    strCell = letter+str(iRow+1)
                    if xCopyStyle:
                        ft = wsOld[strCell].font
                        brdr = wsOld[strCell].border
                        cell = ws[strCell] 
                        cell.border = cell.border.copy(border=brdr.outline,bottom=brdr.bottom,top=brdr.top)
                        cell.font = cell.font.copy(bold=ft.bold)
                    CellContent = (wsOld[strCell].value)
                    if (not CellContent is None) and len(str(CellContent))>0:
                        ws[strCell] = (CellContent)
                
    return ws

def GetMetaData(metafile):
    """GetMetaData
        
    The purpose of :py:func:`.GetMetaData` is to get the meta-data describing the Excel files that should be produced
        
    Parameters
    ----------
    metafile : Path to the file containing the meta-data describing the Excel files that should be produced
    
    Returns
    -------
    dfSITE, dfVAR, lstSampleMethod, lstSampleFilter
    
    """
    
    # ====================================================
    #    CREATE DATA FRAME WITH SITE SPECIFICATIONS
    # ====================================================

    dfSITE = pandas.read_excel(io =metafile, engine="openpyxl",sheet_name ='site',header=0,dtype='str',keep_default_na=False,na_values = {""}).dropna(axis=0,how='all')
    dfSITE.replace(' ', '_', regex=True,inplace=True)
    
    # ====================================================
    #    CREATE DATA FRAME WITH VARIABLE SPECIFICATIONS
    # ====================================================
    
    # Get data
    dfVARIABLES = pandas.read_excel(io =metafile,engine="openpyxl",sheet_name ='variable',header=1,dtype='str',keep_default_na=False,na_values = {""}).dropna(axis=0,how='all')
    lstVAR = dfVARIABLES.columns[1:]

    rowMethodFirst = np.where((dfVARIABLES.iloc[:,0]=='Methods'))[0];
    rowFilterFirst = np.where((dfVARIABLES.iloc[:,0]=='Filters'))[0] ;
    rowUnit = np.where((dfVARIABLES.iloc[:,0]=='Unit'))[0];
    rowMethod = np.arange(rowMethodFirst,rowFilterFirst)
    rowFilter = np.arange(rowFilterFirst,dfVARIABLES.shape[0])

    # individual data frames for units, methods, and filters:
    dfUnit = dfVARIABLES.iloc[rowUnit,:]
    dfMethod = dfVARIABLES.iloc[rowMethod,:]
    dfFilter = dfVARIABLES.iloc[rowFilter,:]

    # Create new data frame with columns for units, methods, and filters; one row corresponds to one variable
    dfVAR = pandas.DataFrame(columns=['unit','method','filter'])
    
    dfVAR['unit'] = dfUnit.iloc[:,1:].values[0]
    dfVAR.index = dfVARIABLES.columns[1:]
    for v in dfVAR.index:
        dfVAR.loc[v,'method']= dfMethod.iloc[:,dfVARIABLES.columns==v].dropna().values[:,0]
        dfVAR.loc[v,'filter']= dfFilter.iloc[:,dfVARIABLES.columns==v].dropna().values[:,0]
    pass

    # ====================================================
    #    CREATE LIST OF FILTERS AVAILABLE FOR SAMPLING
    # ====================================================
    
    Sampling = dfVAR[dfVAR.index=='N/A']
    lstSampleFilter = list(Sampling['filter'][0])
    lstSampleMethod = list(Sampling['method'][0])
    dfVAR = dfVAR.drop(dfVAR.index[dfVAR.index=='N/A'])
    
    
    return dfSITE, dfVAR, lstSampleMethod, lstSampleFilter
    
def Protect(ws,*,nRowCorner,nRowHead,nCol):
    """Protect
        
    The purpose of :py:func:`.Protect` is to update a sheet by protecting its contents
        
    Parameters
    ----------
    ws : Input sheet
    nRowCorner : Number of rows used for project-specific information
    nRowHead : Number of rows used as header
    
    Returns
    -------
    ws : Updated sheet
    
    """
    
    nRow = ws.max_row
    for iRow in range(nRow-nRowCorner-nRowHead):
        for iCol in range(nCol):
            letter = pyxl.utils.get_column_letter(iCol+1)
            cell = ws[letter+str(nRowCorner+nRowHead+iRow+1)]
            cell.protection = pyxl.styles.Protection(locked=False)
    
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.set_password('openpyxl')
    
    return ws